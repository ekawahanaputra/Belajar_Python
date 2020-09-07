# OPENPYXL merupakan sebuah library python yang berfungsi untuk mengolah data file berextensi xlsx
from openpyxl import*     # Mengimport module openpyxl

baca_file = load_workbook(filename = 'SumberData.xlsx')      # load_workbook() berfungsi untuk mengakses data dari file sumber
baca_sheet = baca_file.active                                # untuk mengakses sheet didalam file


# Mengakses cell B2 & lihat hasilnya
print('====SINGLE CELL===')
B2 = baca_sheet.cell(row = 2, column = 2)
print(B2)

# Mengakses nilai cell B2 & lihat hasilnya
B2 = baca_sheet.cell(row = 2, column = 2)
print(B2.value)

# Cara lain mengakses nilai cell B2 & lihat hasilnya
B2 = baca_sheet['B2']
print(B2.value)

# Cara Mengakses Multiple Cell
print('\n====MULTIPLE CELL====')
Cells = baca_sheet['A2','A4']
for a in Cells:
    print(a.value)  XXXXXX